"""
GST Audit Reconciliation Tool
Created by: Purva Doshi
Supports: ITC Reco (GSTR-2B vs Purchase) | Sales Reco (GSTR-1 vs Books)
"""
import streamlit as st
import io, re, datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GST Audit Reco Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)
st.markdown("""
<style>
.hdr{background:#1F3864;padding:1.1rem 1.5rem;border-radius:8px;margin-bottom:1.2rem}
.hdr h1{color:#fff;margin:0;font-size:1.5rem}
.hdr p{color:#AACCEE;margin:3px 0 0;font-size:.85rem}
.credit{color:#7FA8D4;font-size:.78rem;margin-top:2px}
.fmtbox{background:#EEF2F8;border-left:4px solid #1F3864;padding:.7rem 1rem;
        border-radius:0 6px 6px 0;font-size:.82rem;margin:.4rem 0 .8rem}
div[data-testid="stDownloadButton"] button{
    background:#1F3864;color:white;border:none;padding:.6rem 1.2rem;
    border-radius:6px;font-size:1rem;font-weight:600;width:100%}
</style>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL STYLES
# ─────────────────────────────────────────────────────────────────────────────
NAVY="1F3864"; BLUE="2E75B6"; GREEN="375623"; TITLE="203864"
MATCH="D6E4BC"; MIS="FFD7D7"; ONLY2="FFD966"; ONLY1="BDD7EE"
RCM_C="E2EFDA"; CUST_C="FCE4D6"; EXEMPT="F0F0F0"
INR="#,##0.00"; NUM="#,##0"

def F(h):  return PatternFill("solid", start_color=h, end_color=h)
def HF(sz=9, c="FFFFFF"): return Font(name="Arial", bold=True, size=sz, color=c)
def DF(sz=9, c="000000", bold=False): return Font(name="Arial", size=sz, color=c, bold=bold)
def BD():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def CTR(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def LFT(): return Alignment(horizontal="left",   vertical="center")
def RGT(): return Alignment(horizontal="right",  vertical="center")

def safe_float(v):
    try: return float(str(v or "0").replace(",", "").strip())
    except: return 0.0

def clean_gstin(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper().strip())

# ─────────────────────────────────────────────────────────────────────────────
#  UNIVERSAL FILE LOADER  — handles .xls / .xlsx / Tally exports
# ─────────────────────────────────────────────────────────────────────────────
def load_workbook_safe(file_bytes):
    """
    Bulletproof workbook loader.
    Tally often exports .xls files that are actually xlsx (ZIP) format.
    Tries multiple strategies to open the file.
    """
    # Strategy 1: direct BytesIO (works for true xlsx and Tally .xls-as-xlsx)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        return wb
    except Exception:
        pass

    # Strategy 2: try with keep_vba flag
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=True)
        return wb
    except Exception:
        pass

    # Strategy 3: write to temp file and reload
    try:
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        os.unlink(tmp_path)
        return wb
    except Exception as e:
        raise ValueError(
            f"Could not open file: {e}\n\n"
            "Please save the file as .xlsx format:\n"
            "In Excel: File → Save As → Excel Workbook (.xlsx)"
        )


# ─────────────────────────────────────────────────────────────────────────────
#  HEADER DETECTION  — flexible, handles Tally multi-line headers
# ─────────────────────────────────────────────────────────────────────────────
def find_header_row(all_rows, max_search=30):
    """
    Finds the header row by looking for key column names.
    Flexible — handles various Tally and SAP export formats.
    Returns (row_index_0based, headers_list) or raises ValueError.
    """
    HEADER_SIGNALS = [
        # (required_set, optional_set) — row qualifies if ALL required + ANY optional found
        ({"date"}, {"particulars", "party", "vendor", "supplier", "customer"}),
        ({"date"}, {"voucher no", "invoice no", "bill no", "doc no"}),
        ({"particulars"}, {"value", "taxable", "amount", "igst", "cgst"}),
        ({"invoice no"}, {"value", "taxable", "amount", "gstin"}),
        ({"voucher no."}, {"value", "taxable", "particulars"}),
    ]

    for i, row in enumerate(all_rows[:max_search]):
        # Get all non-empty cell values as lowercase strings
        vals = [str(v or "").strip().lower() for v in row]
        vals_joined = " | ".join(vals)

        for required, optional in HEADER_SIGNALS:
            # Check if all required terms appear somewhere in this row
            req_found = all(
                any(req in v for v in vals) for req in required
            )
            # Check if any optional term appears
            opt_found = any(
                any(opt in v for v in vals) for opt in optional
            )
            if req_found and opt_found and sum(1 for v in vals if v) >= 3:
                return i, [str(v or "").strip() for v in row]

    # If still not found, try a looser check — any row with 5+ non-empty cells
    # that contains at least one date-like value in the NEXT row
    for i, row in enumerate(all_rows[:max_search]):
        non_empty = [str(v or "").strip() for v in row if v]
        if len(non_empty) >= 5:
            lower = [v.lower() for v in non_empty]
            has_col_names = any(kw in " ".join(lower) for kw in
                               ["date","gstin","value","taxable","igst","cgst","sgst","invoice","voucher"])
            if has_col_names:
                return i, [str(v or "").strip() for v in row]

    raise ValueError(
        "Could not find header row in the file.\n\n"
        "The file must have a header row with column names like:\n"
        "  Date | Particulars/Party | GSTIN | Value | IGST | CGST | SGST\n\n"
        "Common fix: Export from Tally using:\n"
        "  Gateway of Tally → Purchase/Sales Register → Alt+F1 (Detailed) → Alt+E → Excel"
    )


def find_col(headers, keywords):
    """Find column index by matching keywords to header names."""
    hl = [str(h or "").strip().lower() for h in headers]
    # Exact match first
    for kw in keywords:
        for i, h in enumerate(hl):
            if kw == h: return i
    # Contains match
    for kw in keywords:
        for i, h in enumerate(hl):
            if len(kw) > 3 and kw in h: return i
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  ITC RECO PARSERS
# ─────────────────────────────────────────────────────────────────────────────
def parse_purchase_register(file_bytes):
    """
    Parses Tally Detailed Purchase Register OR any flat purchase Excel.
    Returns DataFrame: Date, Party, GSTIN, VchNo, SupplierInv, Value, IGST, CGST, SGST, Category
    """
    wb  = load_workbook_safe(file_bytes)
    ws  = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    hr_idx, hdrs = find_header_row(all_rows)
    hl = [h.lower() for h in hdrs]

    # Identify tax columns
    igst_cols  = [i for i, h in enumerate(hl) if "igst" in h and "input" in h and "custom" not in h]
    cgst_cols  = [i for i, h in enumerate(hl) if "cgst" in h and "input" in h]
    sgst_cols  = [i for i, h in enumerate(hl) if "sgst" in h and "input" in h]
    cust_igst  = [i for i, h in enumerate(hl) if "igst" in h and "custom" in h]
    has_gstin  = any("gstin" in h for h in hl)

    # Detect format: columnar (Format A) vs ledger sub-row (Format B)
    all_tax_cols = igst_cols + cgst_cols + sgst_cols
    format_a = False
    if has_gstin and all_tax_cols:
        for r in all_rows[hr_idx + 1: hr_idx + 1 + min(150, len(all_rows) - hr_idx - 1)]:
            for c in all_tax_cols:
                if c < len(r) and isinstance(r[c], (int, float)) and r[c] > 0:
                    format_a = True; break
            if format_a: break

    rows = []

    if format_a:
        # ── FORMAT A: one row per invoice, all values in columns ──────────────
        ci = {
            "date":  find_col(hdrs, ["date"]),
            "party": find_col(hdrs, ["particulars", "party", "vendor name", "supplier", "party name"]),
            "vch":   find_col(hdrs, ["voucher no.", "voucher no", "invoice no", "doc no", "bill no"]),
            "sinv":  find_col(hdrs, ["supplier invoice no.", "supplier invoice no", "supp inv no", "ref no"]),
            "sdt":   find_col(hdrs, ["supplier invoice date", "supp inv date"]),
            "gstin": find_col(hdrs, ["gstin/uin", "gstin", "gst no"]),
            "val":   find_col(hdrs, ["value", "taxable value", "basic value", "taxable", "basic"]),
        }

        for row in all_rows[hr_idx + 1:]:
            if not any(row): continue
            dt = row[ci["date"]] if ci["date"] is not None and ci["date"] < len(row) else None
            if not isinstance(dt, (datetime.datetime, datetime.date)): continue

            p = str(row[ci["party"]] or "").strip() if ci["party"] is not None and ci["party"] < len(row) else ""
            if not p or p.lower() in ("grand total", "total", ""): continue

            vch   = str(row[ci["vch"]]  or "").strip() if ci["vch"]  is not None and ci["vch"]  < len(row) else ""
            sinv  = str(row[ci["sinv"]] or "").strip() if ci["sinv"] is not None and ci["sinv"] < len(row) else ""
            gstin = clean_gstin(row[ci["gstin"]]) if ci["gstin"] is not None and ci["gstin"] < len(row) else ""
            val   = safe_float(row[ci["val"]]) if ci["val"] is not None and ci["val"] < len(row) else 0

            igst  = sum(safe_float(row[c]) for c in igst_cols if c < len(row))
            cgst  = sum(safe_float(row[c]) for c in cgst_cols if c < len(row))
            sgst  = sum(safe_float(row[c]) for c in sgst_cols if c < len(row))
            cust  = sum(safe_float(row[c]) for c in cust_igst  if c < len(row))

            cat = ("Custom Duty / Import" if cust > 0
                   else "No ITC / RCM"    if igst == cgst == sgst == 0
                   else "Regular B2B")

            rows.append({"Date": dt, "Party": p, "GSTIN": gstin, "VchNo": vch,
                         "SupplierInv": sinv, "Value": val,
                         "IGST": igst, "CGST": cgst, "SGST": sgst,
                         "TotalTax": igst + cgst + sgst, "Category": cat})

    else:
        # ── FORMAT B: ledger sub-row (older Tally detailed export) ────────────
        cur = {}; ig = cg = sg = tx = 0.0

        def save():
            if cur.get("p") and cur.get("v"):
                cust_amt = cur.get("cust", 0)
                cat = ("Custom Duty / Import" if cust_amt > 0
                       else "No ITC / RCM"    if ig == cg == sg == 0
                       else "Regular B2B")
                rows.append({
                    "Date": cur["d"], "Party": cur["p"], "GSTIN": cur.get("g", ""),
                    "VchNo": cur["v"], "SupplierInv": cur.get("si", ""),
                    "Value": tx, "IGST": ig, "CGST": cg, "SGST": sg,
                    "TotalTax": ig + cg + sg, "Category": cat
                })

        for row in all_rows[hr_idx + 1:]:
            if not any(row): continue
            r  = [row[i] if i < len(row) else None for i in range(max(8, len(row)))]
            dt = r[0]; part = r[1]; vn = r[2]; si = r[3]; db = r[6]

            if isinstance(dt, (datetime.datetime, datetime.date)):
                save()
                p = str(part or "").strip()
                if not p or p.lower() in ("total", "grand total"):
                    cur = {}; ig = cg = sg = tx = 0.0; continue
                cur = {"d": dt, "p": p, "v": str(vn or ""), "si": str(si or ""), "g": ""}
                ig = cg = sg = tx = 0.0
            elif part and cur:
                lbl = str(part).upper()
                amt = safe_float(db)
                if "IGST" in lbl and "CUSTOM" in lbl:
                    cur["cust"] = cur.get("cust", 0) + amt; ig += amt
                elif "IGST" in lbl:  ig += amt
                elif "CGST" in lbl:  cg += amt
                elif "SGST" in lbl or "UTGST" in lbl: sg += amt
                elif "TDS" not in lbl and "ROUNDING" not in lbl and amt > 0: tx += amt
        save()

    if not rows:
        raise ValueError(
            "No invoice data found in Purchase Register.\n\n"
            "Ensure the file has rows with Date, Party/Supplier, and Tax amounts.\n"
            "Export from Tally: Gateway → Purchase Register → Detailed → Alt+E → Excel"
        )

    return pd.DataFrame(rows)


def parse_gstr2b(file_bytes):
    wb = load_workbook_safe(file_bytes)
    if "B2B" not in wb.sheetnames:
        raise ValueError(
            "GSTR-2B file must contain a sheet named 'B2B'.\n"
            "Download Excel format from: GST Portal → Returns → GSTR-2B → Download Excel"
        )
    ws = wb["B2B"]; rows = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row[0] or not str(row[0]).strip(): continue
        rows.append({
            "GSTIN":    clean_gstin(row[0]),
            "TradeNm":  str(row[1] or "").strip(),
            "InvNo":    str(row[2] or "").strip(),
            "Taxable":  safe_float(row[8]),
            "IGST":     safe_float(row[9]),
            "CGST":     safe_float(row[10]),
            "SGST":     safe_float(row[11]),
            "TotalTax": safe_float(row[9]) + safe_float(row[10]) + safe_float(row[11]),
            "RCM":      str(row[7] or "").upper() == "YES",
            "ITCAvail": str(row[15] or "").strip(),
        })
    if not rows:
        raise ValueError("No data found in GSTR-2B B2B sheet.")
    return pd.DataFrame(rows)


def run_itc_reco(pr_bytes, b2b_bytes):
    df_t  = parse_purchase_register(pr_bytes)
    df_2b = parse_gstr2b(b2b_bytes)

    def tkey(r): return r["GSTIN"] if r["GSTIN"] else re.sub(r"[^a-z0-9]", "", r["Party"].lower())
    def gkey(r): return r["GSTIN"] if r["GSTIN"] else re.sub(r"[^a-z0-9]", "", r["TradeNm"].lower())

    df_t["_k"]  = df_t.apply(tkey, axis=1)
    df_2b["_k"] = df_2b.apply(gkey, axis=1)

    ts = df_t.groupby("_k").agg(
        T_Party=("Party","first"), T_GSTIN=("GSTIN","first"), T_Inv=("VchNo","count"),
        T_IGST=("IGST","sum"), T_CGST=("CGST","sum"), T_SGST=("SGST","sum"),
        T_Tax=("TotalTax","sum"), T_Cat=("Category","first"),
    ).reset_index()

    gs = df_2b.groupby("_k").agg(
        G_Party=("TradeNm","first"), G_GSTIN=("GSTIN","first"), G_Inv=("InvNo","count"),
        G_IGST=("IGST","sum"), G_CGST=("CGST","sum"), G_SGST=("SGST","sum"),
        G_Tax=("TotalTax","sum"), G_RCM=("RCM","any"),
    ).reset_index()

    reco = pd.merge(ts, gs, on="_k", how="outer").fillna(0)

    def status(r):
        if "Custom" in str(r.get("T_Cat","")): return "Custom Duty / Import"
        if r.get("G_RCM", False):              return "RCM — Verify Books"
        ti, gi = r["T_Inv"], r["G_Inv"]
        if ti == 0 and gi > 0: return "Only in 2B"
        if gi == 0 and ti > 0: return "Only in Books"
        if (abs(r["T_IGST"]-r["G_IGST"]) < 1 and
            abs(r["T_CGST"]-r["G_CGST"]) < 1 and
            abs(r["T_SGST"]-r["G_SGST"]) < 1):
            return "Matched"
        return "Mismatch"

    def remark(r):
        st = r["Status"]
        if st == "Matched":            return "✓ ITC claimable — books match 2B"
        if st == "RCM — Verify Books": return "RCM — ensure self-invoice raised and tax paid in GSTR-3B"
        if st == "Custom Duty / Import": return "Custom duty/Import — verify IMPG in 2B; ITC available if ICEGATE reflected"
        if st == "Only in 2B":         return "Supplier filed but not booked in Tally — check if purchase entry pending"
        if st == "Only in Books":
            if "Import" in str(r.get("T_Cat","")): return "Import/Custom — may reflect in 2B next month (ICEGATE delay)"
            return "Booked in Tally but supplier not filed GSTR-1 — ITC blocked"
        if st == "Mismatch":
            parts = []
            di = r["T_IGST"] - r["G_IGST"]; dc = r["T_CGST"] - r["G_CGST"]
            ds = r["T_SGST"] - r["G_SGST"]
            if abs(di) > 1: parts.append(f"IGST diff ₹{abs(di):,.0f}")
            if abs(dc) > 1: parts.append(f"CGST diff ₹{abs(dc):,.0f}")
            if abs(ds) > 1: parts.append(f"SGST diff ₹{abs(ds):,.0f}")
            inv_d = int(r["T_Inv"] - r["G_Inv"])
            if inv_d != 0: parts.append(f"Inv count diff {inv_d:+d}")
            hint = ("Check prior-month invoices booked this month" if r["T_Tax"] > r["G_Tax"]
                    else "Check invoices pending in books")
            return ("; ".join(parts) + f" — {hint}") if parts else hint
        return ""

    reco["Status"] = reco.apply(status, axis=1)
    reco["Remark"] = reco.apply(remark, axis=1)
    reco["D_IGST"] = reco["T_IGST"] - reco["G_IGST"]
    reco["D_CGST"] = reco["T_CGST"] - reco["G_CGST"]
    reco["D_Tax"]  = reco["T_Tax"]  - reco["G_Tax"]

    ORDER = {"Matched":0,"Mismatch":1,"RCM — Verify Books":2,
             "Custom Duty / Import":3,"Only in 2B":4,"Only in Books":5}
    reco["_s"] = reco["Status"].map(ORDER).fillna(9)
    return reco.sort_values(["_s","T_Party"]).drop(columns=["_s"]).reset_index(drop=True), df_t, df_2b


# ─────────────────────────────────────────────────────────────────────────────
#  SALES RECO PARSERS
# ─────────────────────────────────────────────────────────────────────────────
TALLY_SALES_SHEETS = [
    "Sales 17-18 Register","GST SALES (Local)","IGST (Sales)",
    "Handling charges","Storage Charges ","SALES TO SEZ (IGST)",
    "STORAGE CHARGES - SEZ UNIT","Sales Exempt","Detention Chgs",
    "SHORTAGE IN TRANSIT","Freight Charges (Income)",
]
EXEMPT_CATS = {"Sales Exempt","Detention Chgs","SHORTAGE IN TRANSIT"}

def smart_val_col(headers):
    """Prefer specific charge column over generic 'value' (for Tally service sheets)."""
    hl = [str(h or "").lower() for h in headers]
    for i, h in enumerate(hl):
        if i > 5 and any(kw in h for kw in [
            "storage charges","handling charges","igst sales","gst sales",
            "sales to sez","shortage","detention","freight charges"
        ]):
            return i
    return find_col(headers, ["value","taxable value","taxable","basic value","basic","amount"])

def parse_sales_books(file_bytes):
    wb    = load_workbook_safe(file_bytes)
    rows  = []
    found_tally = any(s in wb.sheetnames for s in TALLY_SALES_SHEETS)
    sheets = ([s for s in TALLY_SALES_SHEETS if s in wb.sheetnames]
              if found_tally else wb.sheetnames)

    for sname in sheets:
        ws = wb[sname]
        all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        try:
            hr_idx, hdrs = find_header_row(all_rows)
        except ValueError:
            continue

        inv_ci = find_col(hdrs, ["voucher no.","voucher no","invoice no","invoice number","bill no","doc no"])
        pty_ci = find_col(hdrs, ["particulars","party","customer","party name","buyer"])
        dat_ci = find_col(hdrs, ["date","invoice date","bill date","voucher date"])
        val_ci = smart_val_col(hdrs)
        gst_ci = find_col(hdrs, ["gstin/uin","gstin","gst no"])

        if inv_ci is None: continue

        for row in all_rows[hr_idx + 1:]:
            inv = str(row[inv_ci] or "").strip() if inv_ci < len(row) else ""
            if not inv: continue
            p   = str(row[pty_ci] or "").strip() if pty_ci is not None and pty_ci < len(row) else ""
            if p.lower() in ("total","grand total","(cancelled )","cancelled"): continue
            try:   val = float(row[val_ci] or 0) if val_ci is not None and val_ci < len(row) else 0
            except: val = 0
            dt    = row[dat_ci] if dat_ci is not None and dat_ci < len(row) else (row[0] if row else None)
            gstin = clean_gstin(row[gst_ci]) if gst_ci is not None and gst_ci < len(row) else ""
            rows.append({"Date":dt,"Party":p,"GSTIN":gstin,"InvNo":inv,
                         "Taxable":val,"Category":sname.strip()})

    if not rows:
        raise ValueError(
            "No data found in Sales/Books file.\n\n"
            "File needs columns: Invoice No / Voucher No + Taxable Value\n"
            "Export from Tally: Sales Register → Detailed → Alt+E → Excel"
        )

    df = pd.DataFrame(rows)
    df["inv_norm"] = df["InvNo"].apply(lambda s: re.sub(r"[-\s]","",str(s).strip().upper()))
    return df.groupby("inv_norm").agg(
        Date=("Date","first"), Party=("Party","first"), GSTIN=("GSTIN","first"),
        InvNo=("InvNo","first"), Taxable=("Taxable","sum"), Category=("Category","first")
    ).reset_index()

def parse_portal_sales(file_bytes):
    wb = load_workbook_safe(file_bytes)
    if "b2b, sez, de" not in wb.sheetnames:
        raise ValueError(
            "Portal file must have a sheet named 'b2b, sez, de'.\n"
            "Download E-Invoice Excel from: GST Portal → E-Invoice → Download Excel"
        )
    ws   = wb["b2b, sez, de"]; rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if not row[0] or not str(row[0]).strip(): continue
        rows.append({
            "GSTIN":   clean_gstin(row[0]),
            "Party":   str(row[1] or "").strip(),
            "InvNo":   str(row[2] or "").strip(),
            "Date":    str(row[3] or ""),
            "Taxable": safe_float(str(row[11] or "0").replace(",","")),
            "IGST":    safe_float(str(row[12] or "0").replace(",","")),
            "CGST":    safe_float(str(row[13] or "0").replace(",","")),
            "SGST":    safe_float(str(row[14] or "0").replace(",","")),
        })
    if not rows:
        raise ValueError("No data found in 'b2b, sez, de' sheet.")
    df = pd.DataFrame(rows)
    df["inv_norm"] = df["InvNo"].apply(lambda s: re.sub(r"[-\s]","",str(s).strip().upper()))
    return df

def run_sales_reco(portal_bytes, books_bytes):
    df_b = parse_sales_books(books_bytes)
    df_p = parse_portal_sales(portal_bytes)
    reco = pd.merge(df_b, df_p, on="inv_norm", how="outer", suffixes=("_b","_p"))
    reco["B_Tax"] = reco["Taxable_b"].fillna(0)
    reco["P_Tax"] = reco["Taxable_p"].fillna(0)
    reco["Diff"]  = reco["B_Tax"] - reco["P_Tax"]

    def status(r):
        cat = str(r.get("Category",""))
        if cat in EXEMPT_CATS and r["P_Tax"] == 0: return "Exempt / Bond Transfer"
        if r["B_Tax"] == 0 and r["P_Tax"] == 0:   return "Skip"
        if r["B_Tax"] == 0: return "Only in Portal"
        if r["P_Tax"] == 0: return "Only in Books"
        if abs(r["Diff"]) < 1: return "Matched"
        return "Value Mismatch"

    reco["Status"] = reco.apply(status, axis=1)
    return (reco[reco["Status"] != "Skip"]
            .sort_values(["Status","inv_norm"])
            .reset_index(drop=True)), df_b, df_p


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL OUTPUT BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
ITC_BG = {"Matched":MATCH,"Mismatch":MIS,"RCM — Verify Books":RCM_C,
           "Custom Duty / Import":CUST_C,"Only in 2B":ONLY2,"Only in Books":ONLY1}
ITC_FC = {"Matched":"1F6E1F","Mismatch":"C00000","RCM — Verify Books":"375623",
           "Custom Duty / Import":"843C0C","Only in 2B":"7F6000","Only in Books":"004080"}
SAL_BG = {"Matched":MATCH,"Value Mismatch":MIS,"Only in Portal":ONLY2,
           "Only in Books":ONLY1,"Exempt / Bond Transfer":EXEMPT}
SAL_FC = {"Matched":"1F6E1F","Value Mismatch":"C00000","Only in Portal":"7F6000",
           "Only in Books":"004080","Exempt / Bond Transfer":"555555"}
SAL_RM = {
    "Matched":               "Books = Portal ✓",
    "Value Mismatch":        "Same invoice, taxable value differs — verify with client",
    "Only in Portal":        "Filed in GSTR-1 but not found in books — check other ledgers",
    "Only in Books":         "In books but not in portal — verify if filed or exempt",
    "Exempt / Bond Transfer":"Exempt supply / Bond Transfer — not e-invoiced (correct)",
}

def _mhdr(ws, text, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws.cell(1,1); c.value = text
    c.font = Font(name="Arial",bold=True,size=12,color="FFFFFF")
    c.fill = F(TITLE); c.alignment = CTR(); ws.row_dimensions[1].height = 28

def _credit_row(ws, ncols, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row,1); c.value = "Created by: Purva Doshi  |  GST Audit Reconciliation Tool"
    c.font = Font(name="Arial",size=8,italic=True,color="AAAAAA")
    c.alignment = CTR(); ws.row_dimensions[row].height = 14

def _detail_sheet(wb, df, cols, title, hc):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    nc = len(cols)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    h = ws.cell(1,1); h.value = title
    h.font = Font(name="Arial",bold=True,size=11,color="FFFFFF")
    h.fill = F(hc); h.alignment = CTR(); ws.row_dimensions[1].height=22

    for ci,(hdr,w,fld) in enumerate(cols,1):
        c = ws.cell(2,ci); c.value=hdr
        c.font=HF(9); c.fill=F(hc); c.alignment=CTR(); c.border=BD()
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in df.iterrows():
        r = ri + 3
        for ci,(hdr,w,fld) in enumerate(cols,1):
            c = ws.cell(r,ci); c.value = row.get(fld,""); c.border=BD(); c.font=DF(9)
            if fld == "Date": c.number_format="DD-MMM-YY"; c.alignment=CTR()
            elif fld in ("Value","IGST","CGST","SGST","TotalTax","Taxable","Taxable Value"):
                c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws.row_dimensions[r].height = 14
    ws.freeze_panes = "A3"


def build_itc_excel(reco, df_t, df_2b, client, period):
    wb     = openpyxl.Workbook()
    counts = reco["Status"].value_counts().to_dict()

    # ── Summary ────────────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title="Summary"; ws1.sheet_view.showGridLines=False
    _mhdr(ws1, f"{client}  —  GSTR-2B vs Tally ITC Reco  |  {period}", 6)
    _credit_row(ws1, 6)

    for ci,v in enumerate(["","Tally (Books)","GSTR-2B (Portal)","Diff IGST","Diff CGST+SGST","Remark"],1):
        c=ws1.cell(4,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()

    srows = [
        ("✅  Matched",          counts.get("Matched",0),            counts.get("Matched",0),           0, 0,   "ITC fully claimable"),
        ("⚠️  Mismatch",         counts.get("Mismatch",0),           counts.get("Mismatch",0),           "-","-","Investigate before claiming"),
        ("🔵  RCM",              counts.get("RCM — Verify Books",0), 0,                                 0, 0,   "Self-invoice + pay in GSTR-3B"),
        ("🟠  Custom / Import",  counts.get("Custom Duty / Import",0),counts.get("Custom Duty / Import",0),"-","-","Verify IMPG in 2B"),
        ("🟡  Only in 2B",       0,                                  counts.get("Only in 2B",0),         "-","-","Check Tally booking"),
        ("⚪  Only in Books",    counts.get("Only in Books",0),      0,                                 "-","-","Supplier not filed / Import delay"),
    ]
    for ri,rd in enumerate(srows, 5):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            if ci in (2,3,4,5) and isinstance(val,(int,float)): c.number_format="#,##0"; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16

    # ITC amounts
    ws1.merge_cells("A12:F12"); c=ws1.cell(12,1); c.value="ITC AMOUNTS (₹)"
    c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ci,v in enumerate(["","IGST","CGST","SGST","Total Tax",""],1):
        c=ws1.cell(13,ci); c.value=v; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()

    itc_data = [
        ("2B Total",    df_2b["IGST"].sum(), df_2b["CGST"].sum(), df_2b["SGST"].sum(), df_2b["TotalTax"].sum()),
        ("Books Total", df_t["IGST"].sum(),  df_t["CGST"].sum(),  df_t["SGST"].sum(),  df_t["TotalTax"].sum()),
        ("✅ Matched (Claimable)",
         reco[reco.Status=="Matched"]["G_IGST"].sum(),
         reco[reco.Status=="Matched"]["G_CGST"].sum(),
         reco[reco.Status=="Matched"]["G_SGST"].sum(),
         reco[reco.Status=="Matched"]["G_Tax"].sum()),
    ]
    for ri,rd in enumerate(itc_data, 14):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ri==16))
            if ci in (2,3,4,5): c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16
    for col,w in zip("ABCDEF",[32,14,14,14,14,38]): ws1.column_dimensions[col].width=w

    # ── ITC Reco sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("ITC Reco"); ws2.sheet_view.showGridLines=False
    _mhdr(ws2, f"{client}  —  ITC Reco (GSTR-2B vs Tally)  |  {period}", 15)
    _credit_row(ws2, 15)

    rcols = [
        ("#",4),("Party Name",38),("GSTIN",20),
        ("Tally\n# Inv",8),("Tally\nIGST",15),("Tally\nCGST",15),("Tally\nSGST",15),("Tally\nTotal Tax",15),
        ("2B\n# Inv",8),("2B\nIGST",15),("2B\nCGST",15),("2B\nSGST",15),("2B\nTotal Tax",15),
        ("Status",20),("Remarks / Action",52),
    ]
    for ci,(hdr,w) in enumerate(rcols,1):
        c=ws2.cell(4,ci); c.value=hdr; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[4].height=22

    for ri,row in reco.iterrows():
        r=ri+5; st=row["Status"]; bg=ITC_BG.get(st,"FFFFFF")
        party = str(row.get("T_Party") or row.get("G_Party") or "")
        gstin = str(row.get("T_GSTIN") or row.get("G_GSTIN") or "")
        vals  = [ri+1, party, gstin,
                 int(row["T_Inv"]),row["T_IGST"],row["T_CGST"],row["T_SGST"],row["T_Tax"],
                 int(row["G_Inv"]),row["G_IGST"],row["G_CGST"],row["G_SGST"],row["G_Tax"],
                 st, row["Remark"]]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(r,ci); c.value=val; c.fill=F(bg); c.border=BD()
            if ci in (5,6,7,8,10,11,12,13): c.font=DF(9); c.alignment=RGT(); c.number_format=INR
            elif ci in (4,9): c.font=DF(9); c.alignment=CTR(); c.number_format=NUM
            elif ci==14: c.font=Font(name="Arial",bold=True,size=9,color=ITC_FC.get(st,"333")); c.alignment=CTR()
            else: c.font=DF(9); c.alignment=LFT()
        ws2.row_dimensions[r].height=15

    tr = len(reco)+5
    ws2.merge_cells(f"A{tr}:C{tr}"); tc=ws2.cell(tr,1); tc.value="TOTAL"
    tc.font=HF(9); tc.fill=F(NAVY); tc.alignment=CTR(); tc.border=BD()
    for ci in range(4,14):
        c=ws2.cell(tr,ci); c.font=HF(9); c.fill=F(NAVY); c.border=BD()
        if ci in (5,6,7,8,10,11,12,13):
            c.value=f"=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}{tr-1})"
            c.number_format=INR; c.alignment=RGT()
        elif ci in (4,9):
            c.value=f"=SUM({get_column_letter(ci)}5:{get_column_letter(ci)}{tr-1})"
            c.number_format=NUM; c.alignment=CTR()
    ws2.row_dimensions[tr].height=17; ws2.freeze_panes="D5"

    _detail_sheet(wb, df_t, [
        ("Date",12,"Date"),("Party",36,"Party"),("GSTIN",20,"GSTIN"),
        ("Voucher No",20,"VchNo"),("Supplier Inv",20,"SupplierInv"),
        ("Value",15,"Value"),("IGST",14,"IGST"),("CGST",14,"CGST"),("SGST",14,"SGST"),("Category",22,"Category"),
    ], f"Tally Detail | {period}", BLUE)

    df_2b_o = df_2b.rename(columns={"TradeNm":"Trade Name","TotalTax":"Total Tax"})
    _detail_sheet(wb, df_2b_o, [
        ("GSTIN",22,"GSTIN"),("Trade Name",36,"Trade Name"),("Invoice No",20,"InvNo"),
        ("Taxable",15,"Taxable"),("IGST",14,"IGST"),("CGST",14,"CGST"),
        ("SGST",14,"SGST"),("Total Tax",14,"Total Tax"),("ITC Avail",12,"ITCAvail"),("RCM",6,"RCM"),
    ], f"GSTR-2B Detail | {period}", GREEN)

    # Legend
    wl = wb.create_sheet("Legend"); wl.sheet_view.showGridLines=False
    wl.merge_cells("A1:C1"); h=wl.cell(1,1); h.value="Color Legend"
    h.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); h.fill=F(TITLE); h.alignment=CTR()
    for i,(bg,st,desc) in enumerate([
        (MATCH,  "Matched",              "ITC fully claimable"),
        (MIS,    "Mismatch",             "Investigate — difference found"),
        (RCM_C,  "RCM",                 "Self-invoice + pay in GSTR-3B"),
        (CUST_C, "Custom Duty / Import","Verify IMPG in 2B; ICEGATE ITC"),
        (ONLY2,  "Only in 2B",          "Check if purchase to be booked"),
        (ONLY1,  "Only in Books",       "Supplier not filed or import pending"),
    ], 3):
        wl.cell(i,1).value=st; wl.cell(i,1).fill=F(bg); wl.cell(i,1).font=DF(9,bold=True); wl.cell(i,1).border=BD()
        wl.merge_cells(f"B{i}:C{i}"); wl.cell(i,2).value=desc
        wl.cell(i,2).fill=F(bg); wl.cell(i,2).font=DF(9); wl.cell(i,2).border=BD(); wl.cell(i,2).alignment=LFT()
    wl.column_dimensions["A"].width=24; wl.column_dimensions["B"].width=50

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def build_sales_excel(reco, df_b, df_p, client, period):
    wb     = openpyxl.Workbook()
    counts = reco["Status"].value_counts().to_dict()
    b_tot  = df_b["Taxable"].sum(); p_tot = df_p["Taxable"].sum()

    ws1 = wb.active; ws1.title="Summary"; ws1.sheet_view.showGridLines=False
    _mhdr(ws1, f"{client}  —  GSTR-1 vs Books Sales Reco  |  {period}", 5)
    _credit_row(ws1, 5)

    for ci,v in enumerate(["","Books","Portal","Diff","Remark"],1):
        c=ws1.cell(4,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ri,rd in enumerate([
        ("Total Invoices",len(df_b),len(df_p),len(df_b)-len(df_p),""),
        ("✅  Matched",counts.get("Matched",0),counts.get("Matched",0),0,"Clean"),
        ("⚠️  Value Mismatch",counts.get("Value Mismatch",0),counts.get("Value Mismatch",0),0,"Investigate"),
        ("🔵  Only in Books",counts.get("Only in Books",0),0,counts.get("Only in Books",0),"Not filed"),
        ("🟡  Only in Portal",0,counts.get("Only in Portal",0),-counts.get("Only in Portal",0),"Check books"),
        ("⚪  Exempt / Bond",counts.get("Exempt / Bond Transfer",0),0,"N/A","Not e-invoiced (correct)"),
    ],5):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            if ci in (2,3,4) and isinstance(val,(int,float)): c.number_format="#,##0"; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16

    for ci,v in enumerate(["","Books (₹)","Portal (₹)","Diff (₹)",""],1):
        c=ws1.cell(12,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ri,rd in enumerate([
        ("Total Taxable",b_tot,p_tot,b_tot-p_tot),
        ("Matched",reco[reco.Status=="Matched"]["B_Tax"].sum(),reco[reco.Status=="Matched"]["P_Tax"].sum(),0),
        ("Exempt/Bond",reco[reco.Status=="Exempt / Bond Transfer"]["B_Tax"].sum(),0,
         reco[reco.Status=="Exempt / Bond Transfer"]["B_Tax"].sum()),
        ("Only in Books",reco[reco.Status=="Only in Books"]["B_Tax"].sum(),0,
         reco[reco.Status=="Only in Books"]["B_Tax"].sum()),
        ("Only in Portal",0,reco[reco.Status=="Only in Portal"]["P_Tax"].sum(),
         -reco[reco.Status=="Only in Portal"]["P_Tax"].sum()),
    ],13):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            if ci in (2,3,4) and isinstance(val,(int,float)): c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16
    for col,w in zip("ABCDE",[32,18,18,18,28]): ws1.column_dimensions[col].width=w

    ws2 = wb.create_sheet("Invoice Reco"); ws2.sheet_view.showGridLines=False
    _mhdr(ws2, f"{client}  —  GSTR-1 vs Books Invoice Reco  |  {period}", 11)
    _credit_row(ws2, 11)

    rc = [("#",4),("Invoice No (Books)",22),("Invoice No (Portal)",22),("Date",12),
          ("Party",38),("Category",20),("Books (₹)",15),("Portal (₹)",15),
          ("Diff (₹)",14),("Status",24),("Remarks",44)]
    for ci,(hdr,w) in enumerate(rc,1):
        c=ws2.cell(4,ci); c.value=hdr; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[4].height=20

    for ri,row in reco.iterrows():
        r=ri+5; st=row["Status"]; bg=SAL_BG.get(st,"FFFFFF")
        b_inv = str(row.get("InvNo_b") or "")
        p_inv = str(row.get("InvNo")   or row.get("InvNo_p") or "")
        party = str(row.get("Party_b") or row.get("Party_p") or row.get("Party") or "")
        date  = row.get("Date_b") or row.get("Date_p") or row.get("Date") or ""
        cat   = str(row.get("Category",""))
        vals  = [ri+1,b_inv,p_inv,date,party,cat,row["B_Tax"],row["P_Tax"],row["Diff"],st,SAL_RM.get(st,"")]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(r,ci); c.value=val; c.fill=F(bg); c.border=BD()
            if ci in (7,8,9): c.font=DF(9); c.alignment=RGT(); c.number_format=INR
            elif ci==4: c.font=DF(9); c.number_format="DD-MMM-YY"; c.alignment=CTR()
            elif ci==1: c.font=DF(9); c.alignment=CTR()
            elif ci==10: c.font=Font(name="Arial",bold=True,size=9,color=SAL_FC.get(st,"333")); c.alignment=CTR()
            else: c.font=DF(9); c.alignment=LFT()
        ws2.row_dimensions[r].height=14
    ws2.freeze_panes="B5"

    df_b2 = df_b.rename(columns={"InvNo":"Invoice No","Taxable":"Taxable Value"})
    df_p2 = df_p.rename(columns={"InvNo":"Invoice No","Taxable":"Taxable Value"})
    _detail_sheet(wb, df_b2, [
        ("Date",12,"Date"),("Party",40,"Party"),("Invoice No",22,"Invoice No"),
        ("GSTIN",22,"GSTIN"),("Category",22,"Category"),("Taxable Value",16,"Taxable Value"),
    ], f"Books (Sales Register) | {period}", BLUE)
    _detail_sheet(wb, df_p2, [
        ("GSTIN",22,"GSTIN"),("Party",40,"Party"),("Invoice No",22,"Invoice No"),
        ("Date",12,"Date"),("Taxable Value",16,"Taxable Value"),
        ("IGST",14,"IGST"),("CGST",14,"CGST"),("SGST",14,"SGST"),
    ], f"Portal — GSTR-1 | {period}", GREEN)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


# ─────────────────────────────────────────────────────────────────────────────
#  UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hdr">
  <h1>📊 GST Audit Reconciliation Tool</h1>
  <p>ITC Reco (GSTR-2B vs Purchase)  ·  Sales Reco (GSTR-1 vs Books)</p>
  <p class="credit">Created by: Purva Doshi</p>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### Reco Type")
    reco_type = st.radio("", [
        "🔵  ITC Reco  (GSTR-2B vs Purchase)",
        "🟢  Sales Reco  (GSTR-1 vs Books)",
    ], label_visibility="collapsed")
    st.markdown("---")
    client = st.text_input("Client Name", placeholder="e.g. Akin Chemicals Pvt Ltd")
    period = st.text_input("Period",       placeholder="e.g. January 2026")
    st.markdown("---")
    st.markdown("**Legend**")
    st.markdown("🟢 Matched · 🔴 Mismatch  \n🟡 Only Portal/2B · 🔵 Only Books  \n🟠 Custom/Import · ⚪ Exempt")
    st.markdown("---")
    st.caption("Created by: Purva Doshi")

is_itc   = "ITC"   in reco_type
is_sales = "Sales" in reco_type

if is_itc:
    st.markdown("### 📥 ITC Reconciliation — File Upload")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**File 1 — GSTR-2B Excel**")
        st.markdown(
            '<div class="fmtbox">'
            'Portal → Returns → GSTR-2B → Download Excel<br>'
            'Must contain sheet: <code>B2B</code>'
            '</div>', unsafe_allow_html=True)
        f1 = st.file_uploader("GSTR-2B", type=["xlsx","xls"], key="i1", label_visibility="collapsed")
    with c2:
        st.markdown("**File 2 — Tally Purchase Register (Detailed)**")
        st.markdown(
            '<div class="fmtbox">'
            'Gateway of Tally → Purchase Register → <b>Detailed (Alt+F1)</b> → Alt+E → Excel<br>'
            'Required columns: Date · Particulars · GSTIN/UIN · Value · INPUT IGST/CGST/SGST'
            '</div>', unsafe_allow_html=True)
        f2 = st.file_uploader("Purchase Register", type=["xlsx","xls"], key="i2", label_visibility="collapsed")

elif is_sales:
    st.markdown("### 📥 Sales Reconciliation — File Upload")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**File 1 — GSTR-1 / E-Invoice Portal Excel**")
        st.markdown(
            '<div class="fmtbox">'
            'GST Portal → E-Invoice → Download Excel<br>'
            'Must have sheet: <code>b2b, sez, de</code>'
            '</div>', unsafe_allow_html=True)
        f1 = st.file_uploader("GSTR-1 Portal", type=["xlsx","xls"], key="s1", label_visibility="collapsed")
    with c2:
        st.markdown("**File 2 — Tally Sales Register**")
        st.markdown(
            '<div class="fmtbox">'
            'Any Tally sales format accepted<br>'
            'Needs: <code>Voucher No / Invoice No</code> + <code>Value / Taxable Value</code>'
            '</div>', unsafe_allow_html=True)
        f2 = st.file_uploader("Sales Register", type=["xlsx","xls"], key="s2", label_visibility="collapsed")

st.markdown("---")

if f1 and f2:
    if not client.strip(): client = "Client"
    if not period.strip(): period = "Period"

    if st.button("▶  Run Reconciliation", type="primary", use_container_width=True):
        with st.spinner("Running reconciliation..."):
            try:
                b1 = f1.read(); b2 = f2.read()

                if is_itc:
                    reco, df_t, df_2b = run_itc_reco(b2, b1)  # b2=Purchase(File2), b1=2B(File1)
                    cnt = reco["Status"].value_counts().to_dict()
                    buf = build_itc_excel(reco, df_t, df_2b, client, period)
                    fname = f"{client.replace(' ','_')}_ITC_Reco_{period.replace(' ','_')}.xlsx"

                    st.success("✅ ITC Reconciliation complete!")
                    cols = st.columns(6)
                    for col, lbl, key in zip(cols,
                        ["✅ Matched","⚠️ Mismatch","🟡 Only 2B","🔵 Only Books","🔵 RCM","🟠 Custom"],
                        ["Matched","Mismatch","Only in 2B","Only in Books","RCM — Verify Books","Custom Duty / Import"]):
                        col.metric(lbl, cnt.get(key, 0))

                    claimable = reco[reco.Status=="Matched"]["G_Tax"].sum()
                    st.info(f"**ITC Claimable (Matched): ₹{claimable:,.2f}**  ·  "
                            f"Total 2B: ₹{df_2b['TotalTax'].sum():,.2f}  ·  "
                            f"Total Books: ₹{df_t['TotalTax'].sum():,.2f}")

                    disp = reco[["T_Party","G_Party","T_IGST","G_IGST","T_CGST","G_CGST","D_Tax","Status","Remark"]].copy()
                    disp.columns = ["Books Party","2B Party","Books IGST","2B IGST",
                                    "Books CGST","2B CGST","Diff Tax","Status","Remark"]
                    st.dataframe(disp, use_container_width=True, height=380,
                        column_config={
                            "Books IGST": st.column_config.NumberColumn(format="₹ %,.0f"),
                            "2B IGST":    st.column_config.NumberColumn(format="₹ %,.0f"),
                            "Books CGST": st.column_config.NumberColumn(format="₹ %,.0f"),
                            "2B CGST":    st.column_config.NumberColumn(format="₹ %,.0f"),
                            "Diff Tax":   st.column_config.NumberColumn(format="₹ %,.0f"),
                        })

                elif is_sales:
                    reco, df_b, df_p = run_sales_reco(b1, b2)  # b1=Portal(File1), b2=Books(File2)
                    cnt = reco["Status"].value_counts().to_dict()
                    buf = build_sales_excel(reco, df_b, df_p, client, period)
                    fname = f"{client.replace(' ','_')}_Sales_Reco_{period.replace(' ','_')}.xlsx"

                    st.success("✅ Sales Reconciliation complete!")
                    cols = st.columns(5)
                    for col, lbl, key in zip(cols,
                        ["✅ Matched","⚠️ Mismatch","🟡 Portal Only","🔵 Books Only","⚪ Exempt"],
                        ["Matched","Value Mismatch","Only in Portal","Only in Books","Exempt / Bond Transfer"]):
                        col.metric(lbl, cnt.get(key, 0))

                    disp = reco[["inv_norm","B_Tax","P_Tax","Diff","Status"]].copy()
                    disp.columns = ["Invoice No","Books (₹)","Portal (₹)","Diff (₹)","Status"]
                    st.dataframe(disp, use_container_width=True, height=380,
                        column_config={
                            "Books (₹)":  st.column_config.NumberColumn(format="₹ %,.0f"),
                            "Portal (₹)": st.column_config.NumberColumn(format="₹ %,.0f"),
                            "Diff (₹)":   st.column_config.NumberColumn(format="₹ %,.0f"),
                        })

                st.markdown("---")
                st.download_button(
                    "⬇️  Download Excel Reco File",
                    data=buf, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"❌ {e}")
                with st.expander("Technical details (share with developer if needed)"):
                    import traceback
                    st.code(traceback.format_exc())
else:
    st.info("👆 Select reco type in sidebar, upload both files above, then click Run.")
    with st.expander("📋 File format requirements"):
        st.markdown("""
**ITC Reco — Purchase Register:**
Export from Tally: `Gateway → Purchase Register → Alt+F1 (Detailed) → Alt+E → Excel`
One row per invoice. Required columns: `Date · Particulars · GSTIN/UIN · Value · INPUT IGST · INPUT CGST · INPUT SGST`

**SAP / Other ERP:** Any flat Excel with the above column names (exact names don't matter, app auto-detects).

**Sales Reco — Portal file:**
Must have sheet named `b2b, sez, de` (E-Invoice Excel download from GST Portal).

**Sales Reco — Books file:**
Any Tally sales register format. Needs `Voucher No./Invoice No.` + `Value` columns.

**Note:** Both `.xlsx` and `.xls` files are accepted.
        """)
    st.caption("Created by: Purva Doshi  |  GST Audit Reconciliation Tool")
